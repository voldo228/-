import openpyxl

# Открываем файл с данными
workbook = openpyxl.load_workbook("1.xlsx")
worksheet = workbook.active

# Создаем новый Workbook и Worksheet
new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

# Итерируемся по столбцам и строкам и записываем значения в новый Worksheet
for column in range(1, worksheet.max_column + 1):
    for row in range(1, worksheet.max_row + 1):
        # Получаем значение ячейки
        cell_value = worksheet.cell(row=row, column=column).value
        # Записываем значение в новую ячейку с поменяными координатами
        new_worksheet.cell(row=column, column=row).value = cell_value

# Сохраняем новый Workbook в файл
new_workbook.save("2.xlsx")